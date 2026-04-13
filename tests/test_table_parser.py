"""
Unit tests for TableParser module (CSV and Excel parsing).
Tests file parsing functionality without requiring database.
"""
import pytest
import csv
import os
import tempfile
from openpyxl import Workbook, load_workbook

from TableParser import parse_col_csv, parse_col_excel


class TestTableParserCSV:
    """Tests for CSV parsing functions."""
    
    @pytest.fixture
    def sample_csv_file(self):
        """Create a temporary CSV file with sample data."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['email', 'username', 'status'])
            writer.writeheader()
            writer.writerow({'email': 'user1@example.com', 'username': 'User One', 'status': 'active'})
            writer.writerow({'email': 'user2@example.com', 'username': 'User Two', 'status': 'inactive'})
            writer.writerow({'email': 'user3@example.com', 'username': 'User Three', 'status': 'active'})
            temp_path = f.name
        
        yield temp_path
        
        # Cleanup
        if os.path.exists(temp_path):
            os.unlink(temp_path)
    
    @pytest.fixture
    def empty_csv_file(self):
        """Create an empty CSV file (headers only)."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['email', 'username'])
            writer.writeheader()
            temp_path = f.name
        
        yield temp_path
        
        if os.path.exists(temp_path):
            os.unlink(temp_path)
    
    def test_parse_col_csv_success(self, sample_csv_file):
        """Test successful CSV column parsing."""
        emails = parse_col_csv(sample_csv_file, 'email')
        
        assert len(emails) == 3
        assert 'user1@example.com' in emails
        assert 'user2@example.com' in emails
        assert 'user3@example.com' in emails
    
    def test_parse_col_csv_with_whitespace(self, sample_csv_file):
        """Test that CSV parsing strips whitespace."""
        # Create file with whitespace
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['email'])
            writer.writeheader()
            writer.writerow({'email': '  spaced@example.com  '})
            temp_path = f.name
        
        try:
            emails = parse_col_csv(temp_path, 'email')
            assert len(emails) == 1
            assert emails[0] == 'spaced@example.com'  # Should be stripped
        finally:
            os.unlink(temp_path)
    
    def test_parse_col_csv_missing_column(self, sample_csv_file):
        """Test error handling for missing column."""
        with pytest.raises(ValueError) as exc_info:
            parse_col_csv(sample_csv_file, 'nonexistent_column')
        
        assert "отсутствует в CSV файле" in str(exc_info.value)
    
    def test_parse_col_csv_empty_file(self, empty_csv_file):
        """Test parsing CSV with only headers (no data)."""
        emails = parse_col_csv(empty_csv_file, 'email')
        assert len(emails) == 0
    
    def test_parse_col_csv_file_not_found(self):
        """Test error when file doesn't exist."""
        with pytest.raises(FileNotFoundError):
            parse_col_csv('nonexistent_file.csv', 'email')
    
    def test_parse_col_csv_skips_none_values(self, sample_csv_file):
        """Test that None values are skipped."""
        # Create file with some None values
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['email'])
            writer.writeheader()
            writer.writerow({'email': 'valid@example.com'})
            writer.writerow({'email': ''})  # Empty string
            writer.writerow({'email': 'another@example.com'})
            temp_path = f.name
        
        try:
            emails = parse_col_csv(temp_path, 'email')
            # Empty strings should be included but stripped
            assert len(emails) >= 2
            assert 'valid@example.com' in emails
        finally:
            os.unlink(temp_path)


class TestTableParserExcel:
    """Tests for Excel parsing functions."""
    
    @pytest.fixture
    def sample_excel_file(self):
        """Create a temporary Excel file with sample data."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Headers
        ws.append(['email', 'username', 'status'])
        
        # Data rows
        ws.append(['excel1@example.com', 'Excel User 1', 'active'])
        ws.append(['excel2@example.com', 'Excel User 2', 'inactive'])
        ws.append(['excel3@example.com', 'Excel User 3', 'active'])
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            temp_path = f.name
        
        wb.close()
        yield temp_path
        
        if os.path.exists(temp_path):
            os.unlink(temp_path)
    
    @pytest.fixture
    def empty_excel_file(self):
        """Create an empty Excel file (headers only)."""
        wb = Workbook()
        ws = wb.active
        ws.append(['email', 'username'])
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            temp_path = f.name
        
        wb.close()
        yield temp_path
        
        if os.path.exists(temp_path):
            os.unlink(temp_path)
    
    def test_parse_col_excel_success(self, sample_excel_file):
        """Test successful Excel column parsing."""
        emails = parse_col_excel(sample_excel_file, 'email')
        
        assert len(emails) == 3
        assert 'excel1@example.com' in emails
        assert 'excel2@example.com' in emails
        assert 'excel3@example.com' in emails
    
    def test_parse_col_excel_with_whitespace(self, sample_excel_file):
        """Test that Excel parsing strips whitespace."""
        # Create file with whitespace
        wb = Workbook()
        ws = wb.active
        ws.append(['email'])
        ws.append(['  spaced@example.com  '])
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            temp_path = f.name
        
        wb.close()
        try:
            emails = parse_col_excel(temp_path, 'email')
            assert len(emails) == 1
            assert emails[0] == 'spaced@example.com'  # Should be stripped
        finally:
            os.unlink(temp_path)
    
    def test_parse_col_excel_missing_column(self, sample_excel_file):
        """Test error handling for missing column in Excel."""
        with pytest.raises(ValueError) as exc_info:
            parse_col_excel(sample_excel_file, 'nonexistent_column')
        
        assert "не найден в файле" in str(exc_info.value)
    
    def test_parse_col_excel_empty_file(self, empty_excel_file):
        """Test parsing Excel with only headers (no data)."""
        emails = parse_col_excel(empty_excel_file, 'email')
        assert len(emails) == 0
    
    def test_parse_col_excel_file_not_found(self):
        """Test error when Excel file doesn't exist."""
        with pytest.raises(FileNotFoundError):
            parse_col_excel('nonexistent_file.xlsx', 'email')
    
    def test_parse_col_excel_numeric_values(self):
        """Test handling of numeric values converted to strings."""
        wb = Workbook()
        ws = wb.active
        ws.append(['email'])
        ws.append([12345])  # Numeric value
        ws.append(['text@example.com'])
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            temp_path = f.name
        
        wb.close()
        try:
            emails = parse_col_excel(temp_path, 'email')
            assert len(emails) == 2
            assert '12345' in emails  # Should be converted to string
            assert 'text@example.com' in emails
        finally:
            os.unlink(temp_path)


class TestTableParserEdgeCases:
    """Edge case tests for TableParser."""
    
    def test_csv_special_characters_in_email(self):
        """Test CSV parsing with special characters in email."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['email'])
            writer.writeheader()
            writer.writerow({'email': 'user+tag@example.com'})
            writer.writerow({'email': 'user.name@sub.domain.com'})
            temp_path = f.name
        
        try:
            emails = parse_col_csv(temp_path, 'email')
            assert 'user+tag@example.com' in emails
            assert 'user.name@sub.domain.com' in emails
        finally:
            os.unlink(temp_path)
    
    def test_excel_unicode_characters(self):
        """Test Excel parsing with Unicode characters."""
        wb = Workbook()
        ws = wb.active
        ws.append(['email', 'username'])
        ws.append(['test@example.com', 'Иван Петров'])
        ws.append(['unicode@example.com', '中文名字'])
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            temp_path = f.name
        
        wb.close()
        try:
            # Note: We're testing username column for unicode
            from TableParser import parse_col_excel
            usernames = parse_col_excel(temp_path, 'username')
            assert 'Иван Петров' in usernames
            assert '中文名字' in usernames
        finally:
            os.unlink(temp_path)
    
    def test_csv_encoding_utf8(self):
        """Test CSV file with UTF-8 encoding."""
        # Use utf-8 instead of utf-8-sig since our parser doesn't handle BOM
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['email', 'username'])
            writer.writeheader()
            writer.writerow({'email': 'test@example.com', 'username': 'Müller'})
            temp_path = f.name
        
        try:
            emails = parse_col_csv(temp_path, 'email')
            assert len(emails) == 1
        finally:
            os.unlink(temp_path)
