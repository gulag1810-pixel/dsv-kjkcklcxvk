import csv
from openpyxl import load_workbook

def parse_col_csv(file_name, col_name):
    with open(file_name, mode='r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        
        if col_name not in reader.fieldnames:
            raise ValueError(f"Ошибка: Столбец '{col_name}' отсутствует в CSV файле {file_name}")
            
        return [row[col_name].strip() for row in reader if row.get(col_name) is not None]

def parse_col_excel(file_name, col_name):
    wb = load_workbook(file_name, data_only=True)
    sheet = wb.active
    
    headers = [cell.value for cell in sheet[1]] 
    
    if col_name not in headers:
        raise ValueError(f"Ошибка: Столбец '{col_name}' не найден в файле {file_name}")
    
    col_index = headers.index(col_name)
    col_list = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        val = row[col_index]
        if val is not None:
            col_list.append(str(val).strip())
    
    wb.close()
    return col_list




